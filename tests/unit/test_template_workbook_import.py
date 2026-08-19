"""Workbook import tests needed by the Phase 4 starter-suite seed."""

from __future__ import annotations

import io

from openpyxl import Workbook

from app.services import template_service
from app.workers import metadata_jobs
from db import execute

TEMPLATE_HEADERS = [
    "Name", "Model Profile", "Output Profile", "Event Rate Scheme", "Currency",
    "Min Loss Threshold", "Num Max Loss Events", "Franchise Deductible",
    "Unrecognized Occupancy", "Treaty Name Pattern", "Tags",
]
SUITE_HEADERS = [
    "Suite Name", "Position", "Template Name", "Portfolio Name Override",
]


def _workbook(template_rows, suite_rows, *, extra_sheet: str | None = None) -> bytes:
    workbook = Workbook()
    templates = workbook.active
    templates.title = "Templates"
    templates.append(TEMPLATE_HEADERS)
    for row in template_rows:
        templates.append(row)
    suites = workbook.create_sheet("Suites")
    suites.append(SUITE_HEADERS)
    for row in suite_rows:
        suites.append(row)
    if extra_sheet:
        workbook.create_sheet(extra_sheet)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _dlm_row(name="US Wind", scheme="RMS WS"):
    return [
        name, "RMS Default RL25", "RMS Default Output", scheme, "USD",
        1.00, 1, False, "Treat as unknown", None, "US; Wind",
    ]


def test_import_creates_and_updates_templates_and_replaces_suite_items(
    iteration2_db, fake_irp,
):
    metadata_jobs._sync_irp_metadata_body()
    first = _workbook(
        [_dlm_row(), _dlm_row("US Wind Alternate")],
        [["US", 1, "US Wind", None], ["US", 2, "US Wind Alternate", "Alt"]],
    )

    created = template_service.import_template_workbook(
        first, actor_id=iteration2_db.user_a
    )
    second = _workbook(
        [_dlm_row(), _dlm_row("US Wind Alternate")],
        [["US", 1, "US Wind Alternate", "Only"]],
    )
    updated = template_service.import_template_workbook(
        second, actor_id=iteration2_db.user_b
    )

    assert created.applied is True
    assert (created.templates_created, created.suites_created) == (2, 1)
    assert (updated.templates_updated, updated.suites_updated) == (2, 1)
    suite = template_service.list_suites()[0]
    assert [(item["template_name"], item["portfolio_name_override"])
            for item in suite["items"]] == [("US Wind Alternate", "Only")]


def test_import_collects_errors_and_applies_nothing(iteration2_db, fake_irp):
    metadata_jobs._sync_irp_metadata_body()
    source = _workbook(
        [
            _dlm_row(scheme=None),
            _dlm_row(name="US Wind", scheme="RMS WS"),
            ["Bad Types", "RMS Default RL25", "RMS Default Output", "RMS WS", "USD",
             "not a number", 1.5, "sometimes", "unknown", None, None],
        ],
        [
            ["US", 1, "Missing Template", None],
            ["US", 1, "US Wind", None],
            ["US", 2, "US Wind", None],
        ],
    )

    result = template_service.import_template_workbook(source)

    messages = [error.message for error in result.errors]
    assert "Event rate scheme is required for DLM analyses" in messages
    assert any("Duplicate template name" in message for message in messages)
    assert "Min Loss Threshold must be a number" in messages
    assert "Num Max Loss Events must be an integer" in messages
    assert any("Unknown template" in message for message in messages)
    assert any("Duplicate position" in message for message in messages)
    assert any("appears more than once" in message for message in messages)
    assert execute("SELECT id FROM analysis_template", connection="WORKBENCH") == []
    assert execute("SELECT id FROM template_suite", connection="WORKBENCH") == []


def test_import_rejects_unknown_sheet_before_applying(iteration2_db):
    result = template_service.import_template_workbook(
        _workbook([_dlm_row()], [["US", 1, "US Wind", None]], extra_sheet="Notes")
    )

    assert result.errors == (
        template_service.WorkbookError("Notes", None, 'Unknown sheet "Notes"'),
    )
    assert execute("SELECT id FROM analysis_template", connection="WORKBENCH") == []


def test_import_allows_values_absent_from_metadata_cache(iteration2_db):
    source = _workbook(
        [[
            "Unresolved", "Removed Model", "Removed Output", None, "ZZZ",
            1, 1, False, "Skip location during analysis", "Treaty*", "old",
        ]],
        [["Empty", None, None, None], ["Unresolved Suite", 1, "Unresolved", None]],
    )

    result = template_service.import_template_workbook(source)

    assert result.applied is True
    saved = template_service.list_templates()[0]
    assert saved["unresolved"] is True
    assert sorted(suite["item_count"] for suite in template_service.list_suites()) == [0, 1]
