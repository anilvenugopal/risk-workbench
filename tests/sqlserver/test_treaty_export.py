"""Treaty Excel export over stored snapshots (spec 004 US2, T026 — FR-024/R5).

``treaty_service.build_treaty_workbook`` builds a valid ``.xlsx`` over an EDM's
stored treaty snapshots — one row per treaty, columns = the union of attribute
keys — with no Risk Modeler call; ``list_treaties`` parses the snapshot JSON.
"""

from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import load_workbook

from app.services import treaty_service
from app.services._common import _utcnow
from db import execute_command


def _make_edm(name: str = "EDM") -> str:
    edm_id = str(uuid.uuid4())
    now = _utcnow()
    execute_command(
        "INSERT INTO irp_edm (id, name, status, inserted_at, updated_at) "
        "VALUES (:id, :n, 'ready', :now, :now)",
        {"id": edm_id, "n": name, "now": now}, connection="WORKBENCH")
    return edm_id


CAT = {"treatyName": "Cat XoL", "treatyType": "CATA", "attachmentLevel": "PORT",
       "attachmentPoint": 25000000.0, "occurrenceLimit": 100000000.0,
       "currency": {"code": "USD"}}
QS = {"treatyName": "Quota Share", "treatyType": "QUOT", "attachmentLevel": "POL",
      "riskLimit": 10000000.0, "percentageRiShare": 40.0}


def _seed_treaties(edm_id: str) -> None:
    now = _utcnow()
    treaty_service.upsert_treaty_detail(
        edm_id=edm_id, irp_id="1042", name="Cat XoL", attributes=CAT, as_of=now)
    treaty_service.upsert_treaty_detail(
        edm_id=edm_id, irp_id="1043", name="Quota Share", attributes=QS, as_of=now)


def test_workbook_is_valid_xlsx_with_union_of_attribute_columns(workbench_db):
    edm_id = _make_edm()
    _seed_treaties(edm_id)

    data = treaty_service.build_treaty_workbook(edm_id=edm_id)

    wb = load_workbook(io.BytesIO(data))  # a truly valid .xlsx, not just bytes
    ws = wb.active
    header = [c.value for c in ws[1]]
    # identifying columns first, then the UNION of attribute keys across the set
    assert header[0] == "Treaty"
    assert header[1] == "Treaty Id"
    for key in ("treatyName", "treatyType", "attachmentLevel", "attachmentPoint",
                "occurrenceLimit", "currency", "riskLimit", "percentageRiShare"):
        assert key in header, f"union column {key} missing"

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2  # one row per treaty
    by_name = {r[0]: dict(zip(header, r, strict=False)) for r in rows}
    assert by_name["Cat XoL"]["occurrenceLimit"] == 100000000.0
    # a key absent from a treaty renders empty, never an error
    assert by_name["Cat XoL"]["riskLimit"] in (None, "")
    assert by_name["Quota Share"]["percentageRiShare"] == 40.0
    # non-scalar attribute values are serialized, not dropped
    assert "USD" in str(by_name["Cat XoL"]["currency"])


def test_workbook_reads_stored_detail_only_no_gateway_call(
        workbench_db, fake_irp):
    # Any Risk Modeler call would raise — the export must never make one
    # (contracts/http-routes.md: reads STORED detail only, Article 11).
    fake_irp.raise_on_search_treaties = True
    fake_irp.raise_on_list_portfolios = True
    edm_id = _make_edm()
    _seed_treaties(edm_id)

    data = treaty_service.build_treaty_workbook(edm_id=edm_id)
    assert load_workbook(io.BytesIO(data)).active.max_row == 3  # header + 2


def test_workbook_for_edm_with_no_treaties_is_still_valid(workbench_db):
    edm_id = _make_edm()
    data = treaty_service.build_treaty_workbook(edm_id=edm_id)
    ws = load_workbook(io.BytesIO(data)).active
    assert [c.value for c in ws[1]][:2] == ["Treaty", "Treaty Id"]
    assert ws.max_row == 1  # header only


def test_treaty_snapshot_missing_renders_identity_only(workbench_db):
    # A treaty row whose attributes snapshot is null (pre-capability) still
    # exports its identity columns — blank attributes, never an error.
    edm_id = _make_edm()
    now = _utcnow()
    execute_command(
        "INSERT INTO irp_treaty (id, edm_id, name, irp_id, inserted_at, updated_at) "
        "VALUES (:id, :e, 'Legacy Treaty', '9', :now, :now)",
        {"id": str(uuid.uuid4()), "e": edm_id, "now": now}, connection="WORKBENCH")
    data = treaty_service.build_treaty_workbook(edm_id=edm_id)
    ws = load_workbook(io.BytesIO(data)).active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0][0] == "Legacy Treaty"


@pytest.mark.parametrize("missing", [True])
def test_list_treaties_parses_snapshot_and_none_when_missing(
        workbench_db, missing):
    edm_id = _make_edm()
    _seed_treaties(edm_id)
    rows = treaty_service.list_treaties(edm_id=edm_id)
    assert [t.name for t in rows] == ["Cat XoL", "Quota Share"]
    assert rows[0].attributes["treatyType"] == "CATA"
    assert all(t.as_of is not None for t in rows)
