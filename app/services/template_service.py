"""Analysis-template and template-suite persistence and workbook import."""

from __future__ import annotations

import io
import uuid
from collections.abc import Iterable
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from irp_integration.analysis_validation import (
    classify_model_profile,
    validate_analysis_settings,
)
from openpyxl import load_workbook
from sqlalchemy import text

from app.services._common import _txn, _uid, _utcnow
from db import is_unique_violation


@dataclass(frozen=True)
class TemplateValues:
    name: str
    analysis_profile_name: str
    output_profile_name: str
    event_rate_scheme_name: str | None
    currency_code: str
    min_loss_threshold: Decimal = Decimal("1.00")
    num_max_loss_event: int = 1
    franchise_deductible: bool = False
    treat_construction_occupancy_as_unknown: bool = True
    treaty_name_pattern: str | None = None


@dataclass(frozen=True)
class SuiteItemValues:
    template_id: str
    portfolio_name_override: str | None = None


@dataclass(frozen=True)
class WorkbookError:
    sheet: str
    row: int | None
    message: str


@dataclass(frozen=True)
class ImportResult:
    errors: tuple[WorkbookError, ...] = ()
    templates_created: int = 0
    templates_updated: int = 0
    suites_created: int = 0
    suites_updated: int = 0

    @property
    def applied(self) -> bool:
        return not self.errors


class TemplateServiceError(ValueError):
    pass


class TemplateValidationError(TemplateServiceError):
    def __init__(self, errors: Iterable[str]):
        self.errors = tuple(errors)
        super().__init__("; ".join(self.errors))


class TemplateInUseError(TemplateServiceError):
    def __init__(self, suite_names: Iterable[str]):
        self.suite_names = tuple(suite_names)
        super().__init__(
            "Template is used by: " + ", ".join(self.suite_names)
        )


def _clean_optional(value: str | None) -> str | None:
    cleaned = value.strip() if isinstance(value, str) else value
    return cleaned or None


def _template_params(values: TemplateValues) -> dict:
    return {
        "name": values.name.strip(),
        "profile": values.analysis_profile_name.strip(),
        "output": values.output_profile_name.strip(),
        "scheme": _clean_optional(values.event_rate_scheme_name),
        "currency": values.currency_code.strip(),
        "threshold": str(
            Decimal(values.min_loss_threshold).quantize(Decimal("0.01"))
        ),
        "max_events": int(values.num_max_loss_event),
        "franchise": bool(values.franchise_deductible),
        "occupancy": bool(values.treat_construction_occupancy_as_unknown),
        "treaty": _clean_optional(values.treaty_name_pattern),
    }


def _row(conn, sql: str, params: dict | None = None) -> dict | None:
    result = conn.execute(text(sql), params or {}).mappings().first()
    return dict(result) if result is not None else None


def _rows(conn, sql: str, params: dict | None = None) -> list[dict]:
    return [dict(row) for row in conn.execute(text(sql), params or {}).mappings()]


def _profile(conn, name: str) -> dict | None:
    return _row(
        conn,
        """
        SELECT name, is_accumulation, software_version_code, peril_code,
               model_region_code
        FROM irp_model_profile
        WHERE name = :name
        """,
        {"name": name},
    )


def classify_profile(profile_name: str, *, conn=None) -> str | None:
    with _txn(conn) as working:
        profile = _profile(working, profile_name)
        if profile is None:
            return None
        if profile["is_accumulation"]:
            return "Accumulation"
        version = profile["software_version_code"]
        return classify_model_profile(version) if version is not None else None


def _validate_template(conn, params: dict) -> list[str]:
    errors = []
    for label, key in (
        ("Template name", "name"),
        ("Model profile", "profile"),
        ("Output profile", "output"),
        ("Currency", "currency"),
    ):
        if not params[key]:
            errors.append(f"{label} is required")

    profile = _profile(conn, params["profile"]) if params["profile"] else None
    if profile is None or profile["is_accumulation"]:
        return errors
    version = profile["software_version_code"]
    if version is None:
        return errors

    scheme = None
    if params["scheme"]:
        scheme = _row(
            conn,
            """
            SELECT peril_code, model_region_code
            FROM irp_event_rate_scheme
            WHERE name = :name
            """,
            {"name": params["scheme"]},
        )
    pair_known = (
        scheme is not None
        and profile["peril_code"] is not None
        and profile["model_region_code"] is not None
    )
    errors.extend(validate_analysis_settings(
        software_version_code=version,
        scheme_provided=bool(params["scheme"]),
        profile_peril_code=profile["peril_code"] or "",
        profile_model_region_code=profile["model_region_code"] or "",
        scheme_peril_code=scheme["peril_code"] if pair_known else None,
        scheme_model_region_code=(
            scheme["model_region_code"] if pair_known else None
        ),
    ))
    return errors


def _replace_tags(conn, template_id: str, tags: Iterable[str], actor_id: str | None) -> None:
    conn.execute(
        text("DELETE FROM analysis_template_tag WHERE template_id = :id"),
        {"id": template_id},
    )
    seen: set[str] = set()
    for tag in tags:
        name = tag.strip()
        key = name.casefold()
        if not name or key in seen:
            continue
        seen.add(key)
        conn.execute(
            text("""
                INSERT INTO analysis_template_tag
                    (template_id, tag_name, inserted_at, inserted_by)
                VALUES (:template_id, :tag_name, :now, :actor)
            """),
            {
                "template_id": template_id,
                "tag_name": name,
                "now": _utcnow(),
                "actor": actor_id,
            },
        )


def save_template(
    values: TemplateValues,
    *,
    tags: Iterable[str] = (),
    actor_id: str | None = None,
    template_id: str | None = None,
    conn=None,
) -> str:
    params = _template_params(values)
    with _txn(conn) as working:
        errors = _validate_template(working, params)
        if errors:
            raise TemplateValidationError(errors)

        duplicate_sql = """
            SELECT id FROM analysis_template
            WHERE LOWER(name) = LOWER(:name) AND deleted_at IS NULL
        """
        duplicate_params = {"name": params["name"]}
        if template_id is not None:
            duplicate_sql += " AND id <> :id"
            duplicate_params["id"] = template_id
        duplicate = _row(working, duplicate_sql, duplicate_params)
        if duplicate:
            raise TemplateValidationError([
                f'An analysis template named "{params["name"]}" already exists'
            ])

        now = _utcnow()
        saved_id = template_id or str(uuid.uuid4())
        write_params = {
            **params,
            "id": saved_id,
            "actor": actor_id,
            "now": now,
        }
        try:
            with working.begin_nested():
                if template_id is None:
                    working.execute(text("""
                        INSERT INTO analysis_template
                          (id, name, analysis_profile_name, output_profile_name,
                           event_rate_scheme_name, currency_code, min_loss_threshold,
                           num_max_loss_event, franchise_deductible,
                           treat_construction_occupancy_as_unknown, treaty_name_pattern,
                           inserted_at, updated_at, inserted_by, updated_by)
                        VALUES
                          (:id, :name, :profile, :output, :scheme, :currency,
                           :threshold, :max_events, :franchise, :occupancy, :treaty,
                           :now, :now, :actor, :actor)
                    """), write_params)
                else:
                    result = working.execute(text("""
                        UPDATE analysis_template
                        SET name = :name,
                            analysis_profile_name = :profile,
                            output_profile_name = :output,
                            event_rate_scheme_name = :scheme,
                            currency_code = :currency,
                            min_loss_threshold = :threshold,
                            num_max_loss_event = :max_events,
                            franchise_deductible = :franchise,
                            treat_construction_occupancy_as_unknown = :occupancy,
                            treaty_name_pattern = :treaty,
                            updated_at = :now,
                            updated_by = :actor
                        WHERE id = :id AND deleted_at IS NULL
                    """), write_params)
                    if result.rowcount == 0:
                        raise TemplateServiceError("Analysis template was not found")
        except Exception as exc:
            if isinstance(exc, TemplateServiceError):
                raise
            if is_unique_violation(exc):
                raise TemplateValidationError([
                    f'An analysis template named "{params["name"]}" already exists'
                ]) from exc
            raise
        _replace_tags(working, saved_id, tags, actor_id)
        return saved_id


def create_template(
    values: TemplateValues, *, tags: Iterable[str] = (),
    actor_id: str | None = None, conn=None,
) -> str:
    return save_template(values, tags=tags, actor_id=actor_id, conn=conn)


def update_template(
    template_id: str, values: TemplateValues, *, tags: Iterable[str] = (),
    actor_id: str | None = None, conn=None,
) -> str:
    return save_template(
        values, tags=tags, actor_id=actor_id, template_id=template_id, conn=conn
    )


_TEMPLATE_SELECT = """
    SELECT t.*, u.display_name AS author_name,
           mp.id AS model_profile_id, mp.is_accumulation,
           mp.software_version_code,
           op.id AS output_profile_id,
           ers.id AS event_rate_scheme_id,
           c.id AS currency_id
    FROM analysis_template t
    LEFT JOIN app_user u ON u.id = t.inserted_by
    LEFT JOIN irp_model_profile mp ON mp.name = t.analysis_profile_name
    LEFT JOIN irp_output_profile op ON op.name = t.output_profile_name
    LEFT JOIN irp_event_rate_scheme ers ON ers.name = t.event_rate_scheme_name
    LEFT JOIN irp_currency c ON c.code = t.currency_code
"""


def _decorate_template(row: dict, tags: Iterable[str] = ()) -> dict:
    row["id"] = _uid(row["id"])
    row["tags"] = list(tags)
    row["model_profile_unresolved"] = row["model_profile_id"] is None
    row["output_profile_unresolved"] = row["output_profile_id"] is None
    row["event_rate_scheme_unresolved"] = (
        bool(row["event_rate_scheme_name"])
        and row["event_rate_scheme_id"] is None
    )
    row["currency_unresolved"] = row["currency_id"] is None
    row["unresolved"] = any((
        row["model_profile_unresolved"],
        row["output_profile_unresolved"],
        row["event_rate_scheme_unresolved"],
        row["currency_unresolved"],
    ))
    if row["model_profile_id"] is None:
        row["profile_family"] = None
    elif row["is_accumulation"]:
        row["profile_family"] = "Accumulation"
    elif row["software_version_code"] is not None:
        row["profile_family"] = classify_model_profile(
            row["software_version_code"]
        )
    else:
        row["profile_family"] = None
    return row


def get_template(template_id: str, *, conn=None) -> dict | None:
    with _txn(conn) as working:
        row = _row(
            working,
            _TEMPLATE_SELECT + " WHERE t.id = :id AND t.deleted_at IS NULL",
            {"id": template_id},
        )
        if row is None:
            return None
        tags = _rows(
            working,
            """
            SELECT tag_name FROM analysis_template_tag
            WHERE template_id = :id ORDER BY tag_name
            """,
            {"id": template_id},
        )
        return _decorate_template(row, (tag["tag_name"] for tag in tags))


def list_templates(q: str = "", *, conn=None) -> list[dict]:
    with _txn(conn) as working:
        rows = _rows(
            working,
            _TEMPLATE_SELECT + """
            WHERE t.deleted_at IS NULL AND LOWER(t.name) LIKE :q
            ORDER BY t.name
            """,
            {"q": f"%{q.strip().lower()}%"},
        )
        tags_by_id: dict[str, list[str]] = {}
        for tag in _rows(working, """
            SELECT tt.template_id, tt.tag_name
            FROM analysis_template_tag tt
            JOIN analysis_template t ON t.id = tt.template_id
            WHERE t.deleted_at IS NULL
            ORDER BY tt.tag_name
        """):
            tags_by_id.setdefault(_uid(tag["template_id"]), []).append(tag["tag_name"])
        return [
            _decorate_template(row, tags_by_id.get(_uid(row["id"]), []))
            for row in rows
        ]


def list_tag_names(*, conn=None) -> list[str]:
    with _txn(conn) as working:
        return [row["tag_name"] for row in _rows(working, """
            SELECT DISTINCT tt.tag_name
            FROM analysis_template_tag tt
            JOIN analysis_template t ON t.id = tt.template_id
            WHERE t.deleted_at IS NULL
            ORDER BY tt.tag_name
        """)]


def delete_template(
    template_id: str, *, actor_id: str | None = None, conn=None,
) -> None:
    with _txn(conn) as working:
        suites = _rows(working, """
            SELECT DISTINCT s.name
            FROM template_suite_item i
            JOIN template_suite s ON s.id = i.suite_id
            WHERE i.template_id = :id AND s.deleted_at IS NULL
            ORDER BY s.name
        """, {"id": template_id})
        if suites:
            raise TemplateInUseError(row["name"] for row in suites)
        result = working.execute(text("""
            UPDATE analysis_template
            SET deleted_at = :now, updated_at = :now, updated_by = :actor
            WHERE id = :id AND deleted_at IS NULL
        """), {"id": template_id, "now": _utcnow(), "actor": actor_id})
        if result.rowcount == 0:
            raise TemplateServiceError("Analysis template was not found")


def _validate_suite(conn, name: str, items: list[SuiteItemValues]) -> None:
    errors = []
    if not name.strip():
        errors.append("Suite name is required")
    ids = [_uid(item.template_id) for item in items]
    if len(ids) != len(set(ids)):
        errors.append("A template can appear only once in a suite")
    if ids:
        params = {f"id{index}": value for index, value in enumerate(ids)}
        marks = ", ".join(f":id{index}" for index in range(len(ids)))
        found = {
            _uid(row["id"])
            for row in _rows(conn, f"""
                SELECT id FROM analysis_template
                WHERE deleted_at IS NULL AND id IN ({marks})
            """, params)
        }
        missing = [value for value in ids if value not in found]
        if missing:
            errors.append("Every suite entry must reference a live analysis template")
    if errors:
        raise TemplateValidationError(errors)


def save_suite(
    name: str,
    items: Iterable[SuiteItemValues],
    *,
    actor_id: str | None = None,
    suite_id: str | None = None,
    conn=None,
) -> str:
    item_list = list(items)
    clean_name = name.strip()
    with _txn(conn) as working:
        _validate_suite(working, clean_name, item_list)
        duplicate_sql = """
            SELECT id FROM template_suite
            WHERE LOWER(name) = LOWER(:name) AND deleted_at IS NULL
        """
        duplicate_params = {"name": clean_name}
        if suite_id is not None:
            duplicate_sql += " AND id <> :id"
            duplicate_params["id"] = suite_id
        duplicate = _row(working, duplicate_sql, duplicate_params)
        if duplicate:
            raise TemplateValidationError([
                f'A template suite named "{clean_name}" already exists'
            ])

        now = _utcnow()
        saved_id = suite_id or str(uuid.uuid4())
        params = {"id": saved_id, "name": clean_name, "now": now, "actor": actor_id}
        try:
            with working.begin_nested():
                if suite_id is None:
                    working.execute(text("""
                        INSERT INTO template_suite
                            (id, name, inserted_at, updated_at, inserted_by, updated_by)
                        VALUES (:id, :name, :now, :now, :actor, :actor)
                    """), params)
                else:
                    result = working.execute(text("""
                        UPDATE template_suite
                        SET name = :name, updated_at = :now, updated_by = :actor
                        WHERE id = :id AND deleted_at IS NULL
                    """), params)
                    if result.rowcount == 0:
                        raise TemplateServiceError("Template suite was not found")
        except Exception as exc:
            if isinstance(exc, TemplateServiceError):
                raise
            if is_unique_violation(exc):
                raise TemplateValidationError([
                    f'A template suite named "{clean_name}" already exists'
                ]) from exc
            raise

        working.execute(
            text("DELETE FROM template_suite_item WHERE suite_id = :id"),
            {"id": saved_id},
        )
        for position, item in enumerate(item_list, start=1):
            working.execute(text("""
                INSERT INTO template_suite_item
                    (id, suite_id, template_id, position, portfolio_name_override,
                     inserted_at, inserted_by)
                VALUES (:id, :suite, :template, :position, :override, :now, :actor)
            """), {
                "id": str(uuid.uuid4()),
                "suite": saved_id,
                "template": item.template_id,
                "position": position,
                "override": _clean_optional(item.portfolio_name_override),
                "now": now,
                "actor": actor_id,
            })
        return saved_id


def create_suite(
    name: str, items: Iterable[SuiteItemValues], *,
    actor_id: str | None = None, conn=None,
) -> str:
    return save_suite(name, items, actor_id=actor_id, conn=conn)


def update_suite(
    suite_id: str, name: str, items: Iterable[SuiteItemValues], *,
    actor_id: str | None = None, conn=None,
) -> str:
    return save_suite(
        name, items, actor_id=actor_id, suite_id=suite_id, conn=conn
    )


def get_suite(suite_id: str, *, conn=None) -> dict | None:
    with _txn(conn) as working:
        suite = _row(working, """
            SELECT s.*, u.display_name AS author_name
            FROM template_suite s
            LEFT JOIN app_user u ON u.id = s.inserted_by
            WHERE s.id = :id AND s.deleted_at IS NULL
        """, {"id": suite_id})
        if suite is None:
            return None
        suite["id"] = _uid(suite["id"])
        items = _rows(working, """
            SELECT i.id, i.template_id, i.position, i.portfolio_name_override,
                   t.name AS template_name, t.deleted_at AS template_deleted_at,
                   mp.id AS model_profile_id, op.id AS output_profile_id,
                   ers.id AS event_rate_scheme_id, c.id AS currency_id,
                   t.event_rate_scheme_name
            FROM template_suite_item i
            LEFT JOIN analysis_template t ON t.id = i.template_id
            LEFT JOIN irp_model_profile mp ON mp.name = t.analysis_profile_name
            LEFT JOIN irp_output_profile op ON op.name = t.output_profile_name
            LEFT JOIN irp_event_rate_scheme ers ON ers.name = t.event_rate_scheme_name
            LEFT JOIN irp_currency c ON c.code = t.currency_code
            WHERE i.suite_id = :id
            ORDER BY i.position
        """, {"id": suite_id})
        for item in items:
            item["id"] = _uid(item["id"])
            item["template_id"] = _uid(item["template_id"])
            item["unresolved"] = (
                item["template_name"] is None
                or item["template_deleted_at"] is not None
                or item["model_profile_id"] is None
                or item["output_profile_id"] is None
                or item["currency_id"] is None
                or (
                    bool(item["event_rate_scheme_name"])
                    and item["event_rate_scheme_id"] is None
                )
            )
        suite["items"] = items
        suite["item_count"] = len(items)
        suite["unresolved"] = any(item["unresolved"] for item in items)
        return suite


def list_suites(*, conn=None) -> list[dict]:
    with _txn(conn) as working:
        ids = _rows(working, """
            SELECT id FROM template_suite
            WHERE deleted_at IS NULL ORDER BY name
        """)
        return [get_suite(_uid(row["id"]), conn=working) for row in ids]


def delete_suite(
    suite_id: str, *, actor_id: str | None = None, conn=None,
) -> None:
    with _txn(conn) as working:
        result = working.execute(text("""
            UPDATE template_suite
            SET deleted_at = :now, updated_at = :now, updated_by = :actor
            WHERE id = :id AND deleted_at IS NULL
        """), {"id": suite_id, "now": _utcnow(), "actor": actor_id})
        if result.rowcount == 0:
            raise TemplateServiceError("Template suite was not found")


def scheme_options(profile_name: str, *, conn=None) -> list[dict]:
    with _txn(conn) as working:
        profile = _profile(working, profile_name)
        if (
            profile is None
            or profile["peril_code"] is None
            or profile["model_region_code"] is None
        ):
            return []
        options = _rows(working, """
            SELECT name, peril_code, model_region_code, model_version_code, is_hd
            FROM irp_event_rate_scheme
            WHERE peril_code = :peril AND model_region_code = :region
            ORDER BY name
        """, {
            "peril": profile["peril_code"],
            "region": profile["model_region_code"],
        })
        selected = len(options) == 1
        for option in options:
            option["selected"] = selected
        return options


def reference_options(*, conn=None) -> dict[str, list[dict]]:
    with _txn(conn) as working:
        profiles = _rows(working, """
            SELECT name, is_accumulation, software_version_code,
                   peril_code, model_region_code
            FROM irp_model_profile ORDER BY name
        """)
        for profile in profiles:
            if profile["is_accumulation"]:
                profile["family"] = "Accumulation"
            elif profile["software_version_code"] is not None:
                profile["family"] = classify_model_profile(
                    profile["software_version_code"]
                )
            else:
                profile["family"] = None
        return {
            "model_profiles": profiles,
            "output_profiles": _rows(
                working, "SELECT name FROM irp_output_profile ORDER BY name"
            ),
            "currencies": _rows(
                working, "SELECT code, name FROM irp_currency ORDER BY code"
            ),
        }


_TEMPLATE_HEADERS = (
    "Name", "Model Profile", "Output Profile", "Event Rate Scheme", "Currency",
    "Min Loss Threshold", "Num Max Loss Events", "Franchise Deductible",
    "Unrecognized Occupancy", "Treaty Name Pattern", "Tags",
)
_SUITE_HEADERS = (
    "Suite Name", "Position", "Template Name", "Portfolio Name Override",
)


def _open_workbook(source: bytes | bytearray | BinaryIO | str | Path):
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    return load_workbook(source, read_only=True, data_only=True)


def _sheet_headers(sheet) -> tuple[dict[str, int], list[WorkbookError]]:
    expected = _TEMPLATE_HEADERS if sheet.title == "Templates" else _SUITE_HEADERS
    values = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    errors = []
    headers: dict[str, int] = {}
    for index, value in enumerate(values):
        if value is None:
            continue
        if not isinstance(value, str) or value not in expected:
            errors.append(WorkbookError(
                sheet.title, 1, f'Unknown header "{value}"'
            ))
            continue
        if value in headers:
            errors.append(WorkbookError(
                sheet.title, 1, f'Duplicate header "{value}"'
            ))
        headers[value] = index
    for header in expected:
        if header not in headers:
            errors.append(WorkbookError(
                sheet.title, 1, f'Missing required column "{header}"'
            ))
    return headers, errors


def _cell(row: tuple, headers: dict[str, int], name: str):
    index = headers.get(name)
    return row[index] if index is not None and index < len(row) else None


def _text_value(
    value, *, sheet: str, row: int, label: str, required: bool,
    errors: list[WorkbookError],
) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            errors.append(WorkbookError(sheet, row, f"{label} is required"))
        return None
    if not isinstance(value, str):
        errors.append(WorkbookError(sheet, row, f"{label} must be text"))
        return None
    return value.strip()


def _decimal_value(value, row: int, errors: list[WorkbookError]) -> Decimal | None:
    if value is None:
        errors.append(WorkbookError("Templates", row, "Min Loss Threshold is required"))
        return None
    if isinstance(value, bool):
        errors.append(WorkbookError("Templates", row, "Min Loss Threshold must be a number"))
        return None
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError):
        errors.append(WorkbookError("Templates", row, "Min Loss Threshold must be a number"))
        return None
    if parsed.as_tuple().exponent < -2:
        errors.append(WorkbookError(
            "Templates", row, "Min Loss Threshold must have at most 2 decimal places"
        ))
        return None
    return parsed.quantize(Decimal("0.01"))


def _integer_value(value, *, sheet: str, row: int, label: str,
                   errors: list[WorkbookError]) -> int | None:
    if value is None:
        errors.append(WorkbookError(sheet, row, f"{label} is required"))
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        errors.append(WorkbookError(sheet, row, f"{label} must be an integer"))
        return None
    parsed = int(value)
    if parsed != value:
        errors.append(WorkbookError(sheet, row, f"{label} must be an integer"))
        return None
    return parsed


def _boolean_value(value, row: int, errors: list[WorkbookError]) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, str) and value.strip().upper() in {"TRUE", "FALSE"}:
        return value.strip().upper() == "TRUE"
    errors.append(WorkbookError(
        "Templates", row, "Franchise Deductible must be TRUE or FALSE"
    ))
    return None


def _parse_template_rows(sheet, headers, conn, errors):
    parsed: list[tuple[int, TemplateValues, list[str]]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in row):
            continue
        name = _text_value(_cell(row, headers, "Name"), sheet="Templates",
                           row=row_number, label="Name", required=True, errors=errors)
        profile = _text_value(_cell(row, headers, "Model Profile"), sheet="Templates",
                              row=row_number, label="Model Profile", required=True, errors=errors)
        output = _text_value(_cell(row, headers, "Output Profile"), sheet="Templates",
                             row=row_number, label="Output Profile", required=True, errors=errors)
        scheme = _text_value(_cell(row, headers, "Event Rate Scheme"), sheet="Templates",
                             row=row_number, label="Event Rate Scheme", required=False,
                             errors=errors)
        currency = _text_value(_cell(row, headers, "Currency"), sheet="Templates",
                               row=row_number, label="Currency", required=True, errors=errors)
        threshold = _decimal_value(
            _cell(row, headers, "Min Loss Threshold"), row_number, errors
        )
        max_events = _integer_value(
            _cell(row, headers, "Num Max Loss Events"), sheet="Templates",
            row=row_number, label="Num Max Loss Events", errors=errors
        )
        franchise = _boolean_value(
            _cell(row, headers, "Franchise Deductible"), row_number, errors
        )
        occupancy_raw = _cell(row, headers, "Unrecognized Occupancy")
        occupancy = None
        if occupancy_raw == "Treat as unknown":
            occupancy = True
        elif occupancy_raw == "Skip location during analysis":
            occupancy = False
        else:
            errors.append(WorkbookError(
                "Templates", row_number,
                "Unrecognized Occupancy must be Treat as unknown or Skip location during analysis",
            ))
        treaty = _text_value(_cell(row, headers, "Treaty Name Pattern"), sheet="Templates",
                             row=row_number, label="Treaty Name Pattern", required=False,
                             errors=errors)
        tags_raw = _text_value(_cell(row, headers, "Tags"), sheet="Templates",
                               row=row_number, label="Tags", required=False, errors=errors)
        tags = [tag.strip() for tag in (tags_raw or "").split(";") if tag.strip()]
        if name:
            key = name.casefold()
            if key in seen:
                errors.append(WorkbookError(
                    "Templates", row_number, f'Duplicate template name "{name}"'
                ))
            seen.add(key)
        required_parsed = all(value is not None for value in (
            name, profile, output, currency, threshold, max_events, franchise, occupancy
        ))
        if not required_parsed:
            continue
        values = TemplateValues(
            name=name,
            analysis_profile_name=profile,
            output_profile_name=output,
            event_rate_scheme_name=scheme,
            currency_code=currency,
            min_loss_threshold=threshold,
            num_max_loss_event=max_events,
            franchise_deductible=franchise,
            treat_construction_occupancy_as_unknown=occupancy,
            treaty_name_pattern=treaty,
        )
        validation = _validate_template(conn, _template_params(values))
        errors.extend(
            WorkbookError("Templates", row_number, message) for message in validation
        )
        parsed.append((row_number, values, tags))
    return parsed


def _parse_suite_rows(sheet, headers, template_names: set[str], errors):
    suites: dict[str, list[tuple[int, int | None, str | None, str | None]]] = {}
    positions: set[tuple[str, int]] = set()
    templates: set[tuple[str, str]] = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in row):
            continue
        name = _text_value(_cell(row, headers, "Suite Name"), sheet="Suites",
                           row=row_number, label="Suite Name", required=True, errors=errors)
        template_name = _text_value(_cell(row, headers, "Template Name"), sheet="Suites",
                                    row=row_number, label="Template Name", required=False,
                                    errors=errors)
        override = _text_value(_cell(row, headers, "Portfolio Name Override"), sheet="Suites",
                               row=row_number, label="Portfolio Name Override", required=False,
                               errors=errors)
        position_raw = _cell(row, headers, "Position")
        position = None
        if position_raw is not None:
            position = _integer_value(position_raw, sheet="Suites", row=row_number,
                                      label="Position", errors=errors)
            if position is not None and position < 1:
                errors.append(WorkbookError(
                    "Suites", row_number, "Position must be at least 1"
                ))
        if not name:
            continue
        if (position is None) != (template_name is None):
            errors.append(WorkbookError(
                "Suites", row_number,
                "Position and Template Name must both be set, or both be empty for an empty suite",
            ))
        key = name.casefold()
        if position is not None:
            position_key = (key, position)
            if position_key in positions:
                errors.append(WorkbookError(
                    "Suites", row_number,
                    f'Duplicate position {position} in suite "{name}"',
                ))
            positions.add(position_key)
        if template_name is not None:
            template_key = (key, template_name.casefold())
            if template_key in templates:
                errors.append(WorkbookError(
                    "Suites", row_number,
                    f'Template "{template_name}" appears more than once in suite "{name}"',
                ))
            templates.add(template_key)
            if template_name.casefold() not in template_names:
                errors.append(WorkbookError(
                    "Suites", row_number,
                    f'Unknown template "{template_name}"',
                ))
        suites.setdefault(name, []).append((row_number, position, template_name, override))
    for name, rows in suites.items():
        if len(rows) > 1 and any(position is None for _, position, _, _ in rows):
            errors.append(WorkbookError(
                "Suites", next(row for row, position, _, _ in rows if position is None),
                f'Empty-suite row cannot be combined with entries for suite "{name}"',
            ))
    return suites


def import_template_workbook(
    source: bytes | bytearray | BinaryIO | str | Path,
    *,
    actor_id: str | None = None,
    conn=None,
) -> ImportResult:
    try:
        workbook = _open_workbook(source)
    except Exception as exc:  # openpyxl reports several format-specific exception types
        return ImportResult(errors=(WorkbookError(
            "Workbook", None, f"Could not read workbook: {exc}"
        ),))

    errors: list[WorkbookError] = []
    expected_sheets = {"Templates", "Suites"}
    for name in workbook.sheetnames:
        if name not in expected_sheets:
            errors.append(WorkbookError(name, None, f'Unknown sheet "{name}"'))
    for name in sorted(expected_sheets - set(workbook.sheetnames)):
        errors.append(WorkbookError(name, None, f'Missing sheet "{name}"'))
    if errors:
        workbook.close()
        return ImportResult(errors=tuple(errors))

    with _txn(conn) as working:
        template_headers, header_errors = _sheet_headers(workbook["Templates"])
        suite_headers, suite_header_errors = _sheet_headers(workbook["Suites"])
        errors.extend(header_errors)
        errors.extend(suite_header_errors)
        if errors:
            workbook.close()
            return ImportResult(errors=tuple(errors))

        parsed_templates = _parse_template_rows(
            workbook["Templates"], template_headers, working, errors
        )
        existing_templates = _rows(working, """
            SELECT id, name FROM analysis_template WHERE deleted_at IS NULL
        """)
        template_names = {row["name"].casefold() for row in existing_templates}
        template_names.update(values.name.casefold() for _, values, _ in parsed_templates)
        parsed_suites = _parse_suite_rows(
            workbook["Suites"], suite_headers, template_names, errors
        )
        workbook.close()
        if errors:
            return ImportResult(errors=tuple(errors))

        existing_template_ids = {
            row["name"].casefold(): _uid(row["id"]) for row in existing_templates
        }
        templates_created = templates_updated = 0
        for _, values, tags in parsed_templates:
            existing_id = existing_template_ids.get(values.name.casefold())
            saved_id = save_template(
                values, tags=tags, actor_id=actor_id,
                template_id=existing_id, conn=working,
            )
            existing_template_ids[values.name.casefold()] = saved_id
            if existing_id:
                templates_updated += 1
            else:
                templates_created += 1

        existing_suites = {
            row["name"].casefold(): _uid(row["id"])
            for row in _rows(working, """
                SELECT id, name FROM template_suite WHERE deleted_at IS NULL
            """)
        }
        suites_created = suites_updated = 0
        for name, rows in parsed_suites.items():
            suite_items = []
            for _, position, template_name, override in sorted(
                rows, key=lambda value: value[1] or 0
            ):
                if position is None or template_name is None:
                    continue
                suite_items.append(SuiteItemValues(
                    template_id=existing_template_ids[template_name.casefold()],
                    portfolio_name_override=override,
                ))
            existing_id = existing_suites.get(name.casefold())
            saved_id = save_suite(
                name, suite_items, actor_id=actor_id,
                suite_id=existing_id, conn=working,
            )
            existing_suites[name.casefold()] = saved_id
            if existing_id:
                suites_updated += 1
            else:
                suites_created += 1

        return ImportResult(
            templates_created=templates_created,
            templates_updated=templates_updated,
            suites_created=suites_created,
            suites_updated=suites_updated,
        )


__all__ = [
    "ImportResult",
    "SuiteItemValues",
    "TemplateInUseError",
    "TemplateServiceError",
    "TemplateValidationError",
    "TemplateValues",
    "WorkbookError",
    "classify_profile",
    "create_suite",
    "create_template",
    "delete_suite",
    "delete_template",
    "get_suite",
    "get_template",
    "import_template_workbook",
    "list_suites",
    "list_tag_names",
    "list_templates",
    "reference_options",
    "save_suite",
    "save_template",
    "scheme_options",
    "update_suite",
    "update_template",
]
